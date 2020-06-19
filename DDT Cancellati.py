#Copia i file dalla cartella di BER-OFFICE a quella di Knos prendendo i
#percorsi da un file excel e rinominando nel contempo i file con il nome
#specificato nella prima colonna Excel

import openpyxl
import shutil
import os

documento_excel = openpyxl.load_workbook('C:\\Users\\administrator.BERARDI\\Desktop\\elenco ddt.xlsx') #openpyxl.load_workbook('C:\\Users\\tommolini\\Desktop\\Script python\\SCORTE DA RIMEDIARE.xlsx')
print( type(documento_excel))

sheet =documento_excel.get_sheet_by_name('elenco ddt') #documento_excel.get_sheet_by_name('elenco ddt')
#print(sheet['A2'].value)
all_rows = sheet.rows
#row_count = sheet.max_row
pathDestinazione='E:\\DDTCANCELLATI\\' #'C:\\Users\\tommolini\\Desktop\\B'

for row in sheet.iter_rows(min_row=2, max_row=None, min_col=None, max_col=None, values_only=False):
	    pathEFileOrigine = row[1].value
	    testa_coda = os.path.split(pathEFileOrigine)
	    nomeFileOrigine = testa_coda[1]
	    nomeFileDestinazione = row[0].value

	    shutil.copy(pathEFileOrigine, pathDestinazione)
	    os.rename(pathDestinazione+'\\'+nomeFileOrigine, pathDestinazione+'\\'+nomeFileDestinazione)
	    print(pathEFileOrigine + ' -> ' + pathDestinazione+'\\'+nomeFileDestinazione +'\n')
